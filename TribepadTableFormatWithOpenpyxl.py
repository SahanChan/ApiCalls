# -*- coding: utf-8 -*-
"""
Created on Thu Aug 25 11:27:50 2022

@author: Jonathan.Moon
"""

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import formatting, styles

wb = load_workbook("Tribepad Insights Report 257.xlsx")
ws = wb.active


table = Table(displayName="RawDataTable", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

cellref = table
green_fill = PatternFill(start_color='28FF44', end_color='28FF44', fill_type='solid')

rule = formatting.Rule(type='expression')

rule.formula = [f'ISNUMBER(SEARCH("allow", {cellref}))']
rule.dxf = styles.differential.DifferentialStyle(fill=green_fill)

ws.conditional_formatting.add(cellref, rule)


"""for row in table:
    for cell in row:
        if cell == "allow":
            cell.fill = PatternFill("solid",start_color=("28FF44"))
"""
wb.save("Tribepad Insights Report 257Formatted.xlsx")
