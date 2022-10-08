#!/usr/bin/env python3

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


tokenUrl = "https://insights.tribepad.com/api/oauth/Sodexo/access_token"
tokenParams = {
    'client_id': 'insights1131105',
    'client_secret': '',
    'grant_type': 'client_credentials',
    'scope': 'insights_api'
}
response = requests.post(tokenUrl, data=tokenParams)
r = response.json()
token = "Bearer " + r["access_token"]
apiHeader = {'Authorization': token}

reportsURL = "https://insights.tribepad.com/api/oauth/Sodexo/report/list"
reports = requests.post(reportsURL, headers=apiHeader)
output = reports.json()

baseURL = "https://insights.tribepad.com/api/oauth/Sodexo/report/"
reportID = 164
# Weekly Full data capture for Jon
getReportURL = baseURL + str(reportID) + "/retrieve"
outputName = "Tribepad Insights Report " + str(reportID) + ".xlsx"

result = requests.post(getReportURL, headers=apiHeader).json()

columnHeadersRaw = result['column_labels']
columnHeaders = list(columnHeadersRaw.values())
resultData = result['data']

reportDF = pd.DataFrame(resultData)
reportDF.columns = columnHeaders
df = reportDF.set_index('Candidate ID')

df.to_excel(outputName)
print(df)
wb = load_workbook(outputName)
ws = wb.active

table = Table(displayName="RawDataTable", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)
wb.save(outputName)
print("Done")
